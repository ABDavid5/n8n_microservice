"""
Convert a rooms JSON schema into an Excel spreadsheet.

The input JSON is a dict keyed by room ids (e.g. "room_8cdc5811"), each value
holding the session data for one room. Every room becomes one row in the output.

Usage:
    python json_to_excel.py input.json [output.xlsx]
"""

import json
import sys
from pathlib import Path

import pandas as pd


def _join(values, sep="\n"):
    """Join a list into a single cell, dropping empties. Non-lists pass through."""
    if isinstance(values, list):
        return sep.join(str(v) for v in values if v not in (None, ""))
    return values


def compute_ai_questions_missed(room):
    """
    From nudge_adherence_all, take the follow_up_question entries in order and
    re-number them starting at 1 (1st follow_up -> 1, 2nd -> 2, ...).

    Each re-numbered position maps 1:1 onto ai_questions_all (position n ->
    ai_questions_all[n-1]). For every follow_up_question whose status is
    "missed", return the corresponding question text.
    """
    questions = room.get("ai_questions_all") or []
    nudges = room.get("nudge_adherence_all") or []

    # Follow-up nudges in their original order; original ordinals only decide
    # ordering, the re-numbering (1-based) is what indexes into the questions.
    follow_ups = [n for n in nudges if n.get("nudge") == "follow_up_question"]
    follow_ups.sort(key=lambda n: n.get("ordinal", 0))

    missed = []
    for position, nudge in enumerate(follow_ups, start=1):
        if str(nudge.get("status", "")).lower() == "missed":
            idx = position - 1  # 1-based position -> 0-based list index
            if 0 <= idx < len(questions):
                missed.append(questions[idx])
    return _join(missed)


def room_to_row(room):
    extraction = room.get("extraction") or {}
    vitals = extraction.get("vitals") or {}
    referral = room.get("referral") or {}

    return {
        "pharmacy_id": room.get("pharmacy_id"),
        "patient_id": room.get("patient_id"),
        "created_at": room.get("created_at"),
        "transcript": room.get("transcript"),
        "extraction.vitals": json.dumps(vitals, ensure_ascii=False) if vitals else None,
        "extraction.height": vitals.get("height"),
        "extraction.bp": vitals.get("bp"),
        "extraction.weight": vitals.get("weight"),
        "extraction.temp": vitals.get("temp"),
        "extraction.rbs": vitals.get("rbs"),
        "symptoms": _join(extraction.get("symptoms")),
        "provisional_diagnosis": extraction.get("provisional_diagnosis"),
        "ai_questions_all": _join(room.get("ai_questions_all")),
        "ai_questions_missed": compute_ai_questions_missed(room),
        "referral_message": room.get("referral_message"),
        "specialist": referral.get("specialist"),
    }


def convert(input_path, output_path):
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Support either a top-level dict of rooms, or a wrapper holding one.
    if isinstance(data, dict) and not any(
        k.startswith("room_") for k in data
    ):
        # Look one level down for the rooms container.
        for value in data.values():
            if isinstance(value, dict) and any(
                k.startswith("room_") for k in value
            ):
                data = value
                break

    rows = []
    for room_id, room in data.items():
        if not isinstance(room, dict):
            continue
        row = {"room_id": room_id}
        row.update(room_to_row(room))
        rows.append(row)

    columns = [
        "room_id",
        "pharmacy_id",
        "patient_id",
        "created_at",
        "transcript",
        "extraction.vitals",
        "extraction.height",
        "extraction.bp",
        "extraction.weight",
        "extraction.temp",
        "extraction.rbs",
        "symptoms",
        "provisional_diagnosis",
        "ai_questions_all",
        "ai_questions_missed",
        "referral_message",
        "specialist",
    ]

    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(output_path, index=False, engine="openpyxl")

    # Force vital columns to be stored as text so values like "12.5" and
    # "120/80" are never re-interpreted as numbers by Excel.
    text_columns = [
        "extraction.height",
        "extraction.bp",
        "extraction.weight",
        "extraction.temp",
        "extraction.rbs",
    ]
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb.active
    header = {cell.value: cell.column for cell in ws[1]}
    for name in text_columns:
        col = header.get(name)
        if col is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col)
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.number_format = "@"  # text format
    wb.save(output_path)

    print(f"Wrote {len(df)} rooms to {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python json_to_excel.py input.json [output.xlsx]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else str(
        Path(input_path).with_suffix(".xlsx")
    )
    convert(input_path, output_path)
